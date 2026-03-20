"""
DC Manager Pro - Production Backend API
FastAPI + PostgreSQL + JWT Auth
"""

from fastapi import FastAPI, HTTPException, Depends, status
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from contextlib import asynccontextmanager
from typing import Optional, List, Any, Dict
from datetime import datetime, timedelta
from pydantic import BaseModel
import asyncpg
import hashlib
import hmac
import base64
import json
import os
import io
import logging
import urllib.parse
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Logging ──────────────────────────────────────
logging.basicConfig(level=logging.INFO, format='%(asctime)s %(levelname)s %(message)s')
log = logging.getLogger("dcmanager")

# ── Config from env ──────────────────────────────
# Support both DATABASE_URL and individual params
# Individual params are safer (handles special chars in passwords)
DB_HOST = os.getenv("DB_HOST", "db")
DB_PORT = os.getenv("DB_PORT", "5432")
DB_NAME = os.getenv("DB_NAME", "dcmanager")
DB_USER = os.getenv("DB_USER", "dcuser")
DB_PASS = os.getenv("DB_PASS", "dcpassword123")

# Build URL safely - urllib.parse.quote handles special chars like @ # % etc
_safe_pass = urllib.parse.quote(DB_PASS, safe="")
DB_URL = os.getenv("DATABASE_URL") or \
    f"postgresql://{DB_USER}:{_safe_pass}@{DB_HOST}:{DB_PORT}/{DB_NAME}"

SECRET    = os.getenv("JWT_SECRET", "change-this-in-production-use-32-char-secret!")
TOKEN_TTL = int(os.getenv("TOKEN_TTL_HOURS", "24"))
ALLOWED_ORIGINS = os.getenv("ALLOWED_ORIGINS", "*").split(",")

log.info(f"DB target: {DB_HOST}:{DB_PORT}/{DB_NAME} user={DB_USER}")

# ── Simple JWT ────────────────────────────────────
def b64url(data: bytes) -> str:
    return base64.urlsafe_b64encode(data).rstrip(b"=").decode()

def create_token(payload: dict) -> str:
    header = b64url(json.dumps({"alg":"HS256","typ":"JWT"}).encode())
    payload["exp"] = (datetime.utcnow() + timedelta(hours=TOKEN_TTL)).timestamp()
    body   = b64url(json.dumps(payload).encode())
    sig_in = f"{header}.{body}".encode()
    sig    = b64url(hmac.new(SECRET.encode(), sig_in, hashlib.sha256).digest())
    return f"{header}.{body}.{sig}"

def verify_token(token: str) -> dict:
    try:
        parts = token.split(".")
        if len(parts) != 3: raise ValueError("bad format")
        header, body, sig = parts
        sig_in   = f"{header}.{body}".encode()
        expected = b64url(hmac.new(SECRET.encode(), sig_in, hashlib.sha256).digest())
        if not hmac.compare_digest(sig, expected): raise ValueError("bad signature")
        pad  = "=" * (-len(body) % 4)
        data = json.loads(base64.urlsafe_b64decode(body + pad))
        if data.get("exp", 0) < datetime.utcnow().timestamp(): raise ValueError("expired")
        return data
    except Exception as e:
        raise HTTPException(status_code=401, detail=f"Invalid token: {e}")

def hash_pw(pw: str) -> str:
    return hashlib.sha256(pw.encode()).hexdigest()

# ── DB Pool ───────────────────────────────────────
pool: asyncpg.Pool = None

async def db():
    async with pool.acquire() as conn:
        yield conn

# ── Auth ──────────────────────────────────────────
bearer = HTTPBearer(auto_error=False)

async def get_current_user(
    creds: HTTPAuthorizationCredentials = Depends(bearer),
    conn = Depends(db)
):
    if not creds:
        raise HTTPException(status_code=401, detail="Not authenticated")
    payload = verify_token(creds.credentials)
    row = await conn.fetchrow(
        "SELECT * FROM users WHERE id=$1 AND active=true", payload["sub"]
    )
    if not row:
        raise HTTPException(status_code=401, detail="User not found or inactive")
    return dict(row)

def require_write(user=Depends(get_current_user)):
    if user["role"] not in ("admin","superuser"):
        raise HTTPException(status_code=403, detail="Write access required")
    return user

def require_superuser(user=Depends(get_current_user)):
    if user["role"] != "superuser":
        raise HTTPException(status_code=403, detail="Superuser access required")
    return user

# ── App lifespan ──────────────────────────────────
@asynccontextmanager
async def lifespan(app: FastAPI):
    global pool
    import asyncio
    log.info(f"Connecting to database at {DB_HOST}:{DB_PORT}...")
    for attempt in range(15):
        try:
            pool = await asyncpg.create_pool(DB_URL, min_size=2, max_size=10, ssl=False)
            log.info("✓ Database connected successfully")
            break
        except Exception as e:
            log.warning(f"DB connect attempt {attempt+1}/15 failed: {e}")
            await asyncio.sleep(3)
    else:
        raise RuntimeError("Could not connect to database after 15 attempts")
    yield
    await pool.close()

app = FastAPI(title="DC Manager Pro API", version="1.0.0", lifespan=lifespan)
app.add_middleware(
    CORSMiddleware,
    allow_origins=ALLOWED_ORIGINS,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"]
)

# ═══════════════════════════════════════════════════
# MODELS
# ═══════════════════════════════════════════════════
class LoginRequest(BaseModel):
    username: str
    password: str

class UserCreate(BaseModel):
    name: str
    username: str
    password: Optional[str] = None
    role: str = "user"
    email: Optional[str] = None
    dept: Optional[str] = None

class AssetCreate(BaseModel):
    host: str
    type: str = "Server"
    status: str = "Online"
    server_type: Optional[str] = None
    dc: Optional[str] = None
    rack: Optional[str] = None
    ustart: Optional[int] = None
    uheight: int = 1
    ip: Optional[str] = None
    oob: Optional[str] = None
    mac: Optional[str] = None
    vlan: Optional[str] = None
    ipn: Optional[str] = None
    atag: Optional[str] = None
    sn: Optional[str] = None
    notes: Optional[str] = None
    hw_data: Dict[str, Any] = {}

class RackCreate(BaseModel):
    rack_id: str
    dc: Optional[str] = None
    zone: Optional[str] = None
    row_label: Optional[str] = None
    total_u: int = 42
    notes: Optional[str] = None

class StockCreate(BaseModel):
    cat: str
    brand: Optional[str] = None
    model: Optional[str] = None
    spec: Optional[str] = None
    ff: Optional[str] = None
    ifc: Optional[str] = None
    total: int = 0
    avail: int = 0
    cost: Optional[float] = None
    loc: Optional[str] = None
    notes: Optional[str] = None

class StockTransaction(BaseModel):
    stock_id:     str
    tx_type:      str
    qty:          int
    ref:          Optional[str] = None
    notes:        Optional[str] = None
    allocated_to: Optional[str] = None   # asset hostname — for ALLOCATE/RETURN tracking

class ConnCreate(BaseModel):
    src_host: Optional[str] = None
    src_slot: Optional[str] = None
    src_port: Optional[str] = None
    src_label: Optional[str] = None
    liu_a_rack: Optional[str] = None
    liu_a_host: Optional[str] = None
    liu_a_port: Optional[str] = None
    liu_b_rack: Optional[str] = None
    liu_b_host: Optional[str] = None
    liu_b_port: Optional[str] = None
    dst_host: Optional[str] = None
    dst_port: Optional[str] = None
    cable: Optional[str] = None
    speed: Optional[str] = None
    vlan: Optional[str] = None
    purpose: Optional[str] = None
    notes: Optional[str] = None

class HWField(BaseModel):
    key: str
    label: str
    field_type: str = "text"
    placeholder: Optional[str] = None
    options: Optional[str] = None
    required: bool = False
    sort_order: int = 100

# ═══════════════════════════════════════════════════
# HELPERS
# ═══════════════════════════════════════════════════
def _out(r):
    """Convert asyncpg record to JSON-safe dict"""
    if r is None: return None
    d = dict(r)
    for k, v in d.items():
        if hasattr(v, 'isoformat'): d[k] = v.isoformat()
        elif isinstance(v, asyncpg.pgproto.pgproto.UUID): d[k] = str(v)
        elif not isinstance(v, (str,int,float,bool,dict,list,type(None))): d[k] = str(v)
    if 'hw_data' in d and isinstance(d['hw_data'], str):
        try: d['hw_data'] = json.loads(d['hw_data'])
        except: d['hw_data'] = {}
    return d


async def log_audit(conn, user: dict, action: str, entity: str, entity_id: str,
                    detail: str = None, related_entity: str = None, related_entity_id: str = None):
    """Write an audit log entry.
    related_entity / related_entity_id link this entry to a secondary entity,
    e.g. a stock transaction linked to the asset it was allocated to."""
    try:
        await conn.execute(
            """INSERT INTO audit_log
               (user_id, username, action, entity, entity_id, detail, related_entity, related_entity_id)
               VALUES ($1, $2, $3, $4, $5, $6, $7, $8)""",
            user.get("id"), user.get("username"), action, entity, str(entity_id),
            detail, related_entity, related_entity_id
        )
    except Exception as e:
        import logging
        logging.getLogger("dcmanager").warning(f"Audit log failed: {e}")

# ═══════════════════════════════════════════════════
# HEALTH
# ═══════════════════════════════════════════════════
@app.get("/api/health")
async def health():
    return {"status": "ok", "time": datetime.utcnow().isoformat()}

@app.get("/api/health/db")
async def health_db(conn=Depends(db)):
    await conn.fetchval("SELECT 1")
    return {"status": "ok", "db": "connected"}

# ═══════════════════════════════════════════════════
# AUTH
# ═══════════════════════════════════════════════════
@app.post("/api/auth/login")
async def login(req: LoginRequest, conn=Depends(db)):
    row = await conn.fetchrow(
        "SELECT * FROM users WHERE username=$1 AND pw_hash=$2 AND active=true",
        req.username.lower(), hash_pw(req.password)
    )
    if not row:
        raise HTTPException(status_code=401, detail="Invalid credentials")
    token = create_token({"sub": str(row["id"]), "role": row["role"], "name": row["name"]})
    await conn.execute("UPDATE users SET last_login=NOW() WHERE id=$1", row["id"])
    return {"token": token, "user": {
        "id": str(row["id"]), "name": row["name"],
        "username": row["username"], "role": row["role"],
        "email": row["email"], "dept": row["dept"]
    }}

@app.get("/api/auth/me")
async def me(user=Depends(get_current_user)):
    return {k:v for k,v in user.items() if k != "pw_hash"}

# ═══════════════════════════════════════════════════
# STATS
# ═══════════════════════════════════════════════════
@app.get("/api/stats")
async def stats(conn=Depends(db), _=Depends(get_current_user)):
    assets     = await conn.fetchval("SELECT COUNT(*) FROM assets")
    online     = await conn.fetchval("SELECT COUNT(*) FROM assets WHERE status='Online'")
    racks      = await conn.fetchval("""
        SELECT COUNT(DISTINCT rack_id) FROM (
            SELECT rack_id FROM racks
            UNION
            SELECT rack_id FROM assets WHERE rack_id IS NOT NULL AND rack_id != ''
        ) r""")
    conns      = await conn.fetchval("SELECT COUNT(*) FROM connectivity")
    stock_skus = await conn.fetchval("SELECT COUNT(*) FROM stock")
    low_stock  = await conn.fetchval("SELECT COUNT(*) FROM stock WHERE avail_qty <= 5")
    by_type    = await conn.fetch("SELECT asset_type, COUNT(*) as cnt FROM assets GROUP BY asset_type ORDER BY cnt DESC")
    by_status  = await conn.fetch("SELECT status, COUNT(*) as cnt FROM assets GROUP BY status")
    return {
        "assets": assets, "online": online, "racks": racks,
        "conns": conns, "stock_skus": stock_skus, "low_stock": low_stock,
        "by_type": [dict(r) for r in by_type],
        "by_status": [dict(r) for r in by_status],
    }

# ═══════════════════════════════════════════════════
# ASSETS
# ═══════════════════════════════════════════════════
@app.get("/api/assets")
async def list_assets(q:Optional[str]=None, type:Optional[str]=None,
                      status:Optional[str]=None, conn=Depends(db), _=Depends(get_current_user)):
    where, params = [], []
    if q:
        params.append(f"%{q}%")
        where.append(f"(hostname ILIKE ${len(params)} OR mgmt_ip ILIKE ${len(params)} OR serial_number ILIKE ${len(params)} OR rack_id ILIKE ${len(params)})")
    if type:   params.append(type);   where.append(f"asset_type=${len(params)}")
    if status: params.append(status); where.append(f"status=${len(params)}")
    sql = "SELECT * FROM assets" + (" WHERE "+" AND ".join(where) if where else "") + " ORDER BY hostname"
    rows = await conn.fetch(sql, *params)
    return [_out(r) for r in rows]


@app.get("/api/assets/check-slot")
async def check_slot(rack:str, ustart:int, uheight:int=1, exclude_id:Optional[str]=None,
                     conn=Depends(db), _=Depends(get_current_user)):
    """Return any assets occupying the given rack+U range, optionally excluding one asset id."""
    u_end = ustart + uheight - 1
    q = """SELECT id,hostname,asset_type,u_start,u_height FROM assets
           WHERE rack_id=$1
             AND u_start IS NOT NULL
             AND (u_start <= $3 AND u_start + u_height - 1 >= $2)"""
    params = [rack, ustart, u_end]
    if exclude_id:
        q += " AND id != $4"
        params.append(exclude_id)
    rows = await conn.fetch(q, *params)
    return [_out(r) for r in rows]

@app.get("/api/assets/{aid}")
async def get_asset(aid:str, conn=Depends(db), _=Depends(get_current_user)):
    row = await conn.fetchrow("SELECT * FROM assets WHERE id=$1", aid)
    if not row: raise HTTPException(404,"Asset not found")
    return _out(row)

@app.post("/api/assets", status_code=201)
async def create_asset(a:AssetCreate, conn=Depends(db), user=Depends(require_write)):
    row = await conn.fetchrow("""
        INSERT INTO assets (hostname,asset_type,status,server_type,datacenter,rack_id,u_start,u_height,
            mgmt_ip,oob_ip,mac_addr,vlan,extra_ips,asset_tag,serial_number,notes,hw_data)
        VALUES ($1,$2,$3,$4,$5,$6,$7,$8,$9,$10,$11,$12,$13,$14,$15,$16,$17) RETURNING *""",
        a.host,a.type,a.status,a.server_type,a.dc,a.rack,a.ustart,a.uheight,
        a.ip,a.oob,a.mac,a.vlan,a.ipn,a.atag,a.sn,a.notes,json.dumps(a.hw_data))
    detail = f"Created asset type={a.type} status={a.status} dc={a.dc} rack={a.rack} u={a.ustart}"
    await log_audit(conn, user, "ASSET_CREATED", "asset", str(row["id"]), detail)
    return _out(row)

@app.put("/api/assets/{aid}")
async def update_asset(aid:str, a:AssetCreate, conn=Depends(db), user=Depends(require_write)):
    old = await conn.fetchrow("SELECT * FROM assets WHERE id=$1", aid)
    if not old: raise HTTPException(404,"Asset not found")
    row = await conn.fetchrow("""
        UPDATE assets SET hostname=$1,asset_type=$2,status=$3,server_type=$4,datacenter=$5,
            rack_id=$6,u_start=$7,u_height=$8,mgmt_ip=$9,oob_ip=$10,mac_addr=$11,vlan=$12,
            extra_ips=$13,asset_tag=$14,serial_number=$15,notes=$16,hw_data=$17,updated_at=NOW()
        WHERE id=$18 RETURNING *""",
        a.host,a.type,a.status,a.server_type,a.dc,a.rack,a.ustart,a.uheight,
        a.ip,a.oob,a.mac,a.vlan,a.ipn,a.atag,a.sn,a.notes,json.dumps(a.hw_data),aid)
    # Detect meaningful changes for audit trail
    changes = []
    if old["status"] != a.status:
        changes.append(f"Status: {old['status']} → {a.status}")
    if old["rack_id"] != a.rack:
        changes.append(f"Rack: {old['rack_id'] or '—'} → {a.rack or '—'}")
    if old["datacenter"] != a.dc:
        changes.append(f"DC: {old['datacenter'] or '—'} → {a.dc or '—'}")
    if old["u_start"] != a.ustart:
        changes.append(f"U Position: {old['u_start'] or '—'} → {a.ustart or '—'}")
    if old["asset_type"] != a.type:
        changes.append(f"Type: {old['asset_type']} → {a.type}")
    # Detect hw_data changes (disks etc)
    try:
        old_hw = json.loads(old["hw_data"]) if old["hw_data"] else {}
        new_hw = a.hw_data or {}
        for k in set(list(old_hw.keys()) + list(new_hw.keys())):
            ov, nv = old_hw.get(k), new_hw.get(k)
            if ov != nv and k not in ("po_number","eol_date","app_owner"):
                changes.append(f"HW {k}: {ov or '—'} → {nv or '—'}")
    except: pass
    if not changes:
        changes.append("Asset updated (no structural changes)")
    action = "ASSET_RELOCATED" if any("Rack:" in c or "DC:" in c or "U Position:" in c for c in changes) else              "ASSET_STATUS_CHANGE" if any("Status:" in c for c in changes) else "ASSET_UPDATED"
    await log_audit(conn, user, action, "asset", aid, " | ".join(changes))
    return _out(row)


@app.get("/api/assets/{aid}/history")
async def asset_history(aid:str, conn=Depends(db), _=Depends(get_current_user)):
    rows = await conn.fetch(
        """SELECT * FROM audit_log WHERE entity='asset' AND entity_id=$1
           ORDER BY created_at DESC LIMIT 200""", aid)
    return [_out(r) for r in rows]

@app.get("/api/assets/{aid}/full-history")
async def asset_full_history(aid:str, conn=Depends(db), _=Depends(get_current_user)):
    """Return a unified timeline for an asset:
    - All direct asset audit events (rack moves, status changes, etc.)
    - All stock transactions allocated to this asset (parts usage)
    - All connectivity events where this asset is src or dst
    Sorted newest-first."""
    asset = await conn.fetchrow("SELECT hostname FROM assets WHERE id=$1", aid)
    if not asset:
        raise HTTPException(404, "Asset not found")
    hostname = asset["hostname"]

    # 1. Direct asset audit entries
    asset_rows = await conn.fetch(
        """SELECT *, 'asset_event' AS source_type
           FROM audit_log
           WHERE entity='asset' AND entity_id=$1
           ORDER BY created_at DESC LIMIT 300""", aid)

    # 2. Stock/connectivity events linked to this asset via related_entity_id
    linked_rows = await conn.fetch(
        """SELECT *, 'linked_event' AS source_type
           FROM audit_log
           WHERE related_entity='asset' AND related_entity_id=$1
             AND entity!='asset'
           ORDER BY created_at DESC LIMIT 200""", hostname)

    # 3. Connectivity events where src or dst hostname matches
    conn_rows = await conn.fetch(
        """SELECT *, 'conn_event' AS source_type
           FROM audit_log
           WHERE entity='connectivity'
             AND (related_entity='asset' AND related_entity_id=$1)
           ORDER BY created_at DESC LIMIT 100""", hostname)

    # Merge and sort all rows by created_at descending
    all_rows = list(asset_rows) + list(linked_rows)
    # Deduplicate by id
    seen = set()
    deduped = []
    for r in all_rows:
        rid = str(r["id"])
        if rid not in seen:
            seen.add(rid)
            deduped.append(_out(r))
    deduped.sort(key=lambda x: x.get("created_at",""), reverse=True)
    return deduped[:300]

@app.get("/api/stock/{sid}/history")
async def stock_history(sid:str, conn=Depends(db), _=Depends(get_current_user)):
    """Return full audit history for a stock item — includes all transactions
    and any create/update/delete audit entries."""
    # Audit entries
    audit_rows = await conn.fetch(
        """SELECT *, 'audit' AS source_type FROM audit_log
           WHERE entity='stock' AND entity_id=$1
           ORDER BY created_at DESC LIMIT 200""", sid)
    # Transaction records with richer detail
    tx_rows = await conn.fetch(
        """SELECT *, 'transaction' AS source_type FROM stock_transactions
           WHERE stock_id=$1 ORDER BY created_at DESC LIMIT 200""", sid)

    # Merge, deduplicate, sort
    result = []
    seen_ids = set()
    for r in audit_rows:
        rid = str(r["id"])
        if rid not in seen_ids:
            seen_ids.add(rid)
            d = _out(r)
            d["source_type"] = "audit"
            result.append(d)
    for r in tx_rows:
        d = _out(r)
        d["source_type"] = "transaction"
        result.append(d)
    result.sort(key=lambda x: x.get("created_at",""), reverse=True)
    return result[:300]

@app.get("/api/connectivity/{cid}/history")
async def conn_history(cid:str, conn=Depends(db), _=Depends(get_current_user)):
    """Return audit history for a specific connectivity record."""
    rows = await conn.fetch(
        """SELECT * FROM audit_log
           WHERE entity='connectivity' AND entity_id=$1
           ORDER BY created_at DESC LIMIT 100""", cid)
    return [_out(r) for r in rows]

@app.get("/api/audit")
async def audit_log_list(
    entity:   Optional[str] = None,
    action:   Optional[str] = None,
    username: Optional[str] = None,
    q:        Optional[str] = None,
    since:    Optional[str] = None,   # ISO date string e.g. 2026-03-01
    limit:    int = 200,
    offset:   int = 0,
    conn=Depends(db), user=Depends(get_current_user)
):
    # Both admin and superuser can see the audit log
    if user["role"] not in ("admin", "superuser"):
        raise HTTPException(status_code=403, detail="Admin access required")

    where, params = [], []

    if entity:
        params.append(entity)
        where.append(f"entity=${len(params)}")
    if action:
        params.append(action)
        where.append(f"action=${len(params)}")
    if username:
        params.append(f"%{username}%")
        where.append(f"username ILIKE ${len(params)}")
    if q:
        params.append(f"%{q}%")
        where.append(f"(detail ILIKE ${len(params)} OR entity_id ILIKE ${len(params)} OR related_entity_id ILIKE ${len(params)})")
    if since:
        params.append(since)
        where.append(f"created_at >= ${len(params)}::timestamptz")

    w_sql = (" WHERE " + " AND ".join(where)) if where else ""

    # Total count for pagination
    count_sql = f"SELECT COUNT(*) FROM audit_log{w_sql}"
    total = await conn.fetchval(count_sql, *params)

    # Data
    params.append(limit)
    params.append(offset)
    rows = await conn.fetch(
        f"SELECT * FROM audit_log{w_sql} ORDER BY created_at DESC "
        f"LIMIT ${len(params)-1} OFFSET ${len(params)}",
        *params
    )
    return {
        "total":  total,
        "limit":  limit,
        "offset": offset,
        "items":  [_out(r) for r in rows]
    }

@app.delete("/api/assets/{aid}", status_code=204)
async def delete_asset(aid:str, conn=Depends(db), user=Depends(require_write)):
    old = await conn.fetchrow("SELECT hostname,asset_type,rack_id,datacenter FROM assets WHERE id=$1", aid)
    await conn.execute("DELETE FROM assets WHERE id=$1", aid)
    if old:
        await log_audit(conn, user, "ASSET_DELETED", "asset", aid,
            f"Deleted {old['hostname']} type={old['asset_type']} rack={old['rack_id']} dc={old['datacenter']}")

# ═══════════════════════════════════════════════════
# RACKS
# ═══════════════════════════════════════════════════
@app.get("/api/racks")
async def list_racks(conn=Depends(db), _=Depends(get_current_user)):
    rows = await conn.fetch("SELECT * FROM racks ORDER BY datacenter,zone,row_label,rack_id")
    return [_out(r) for r in rows]

@app.post("/api/racks", status_code=201)
async def create_rack(r:RackCreate, conn=Depends(db), user=Depends(require_write)):
    row = await conn.fetchrow("""
        INSERT INTO racks (rack_id,datacenter,zone,row_label,total_u,notes)
        VALUES ($1,$2,$3,$4,$5,$6) ON CONFLICT (rack_id) DO UPDATE
        SET datacenter=$2,zone=$3,row_label=$4,total_u=$5,notes=$6 RETURNING *""",
        r.rack_id,r.dc,r.zone,r.row_label,r.total_u,r.notes)
    await log_audit(conn, user, "RACK_CREATED", "rack", r.rack_id,
        f"Rack {r.rack_id} dc={r.dc} zone={r.zone} row={r.row_label} u={r.total_u}")
    return _out(row)

@app.put("/api/racks/{rid}")
async def update_rack(rid:str, r:RackCreate, conn=Depends(db), user=Depends(require_write)):
    row = await conn.fetchrow("""
        UPDATE racks SET datacenter=$1,zone=$2,row_label=$3,total_u=$4,notes=$5
        WHERE rack_id=$6 RETURNING *""",
        r.dc,r.zone,r.row_label,r.total_u,r.notes,rid)
    if not row: raise HTTPException(404,"Rack not found")
    await log_audit(conn, user, "RACK_UPDATED", "rack", rid,
        f"Updated: dc={r.dc} zone={r.zone} row={r.row_label} u={r.total_u}")
    return _out(row)

@app.delete("/api/racks/{rid}", status_code=204)
async def delete_rack(rid:str, conn=Depends(db), user=Depends(require_superuser)):
    # Check if rack has assets
    count = await conn.fetchval("SELECT COUNT(*) FROM assets WHERE rack_id=$1", rid)
    if count > 0:
        raise HTTPException(400, f"Cannot delete rack '{rid}' — it has {count} asset(s). Move them first.")
    await conn.execute("DELETE FROM racks WHERE rack_id=$1", rid)
    await log_audit(conn, user, "RACK_DELETED", "rack", rid, f"Rack {rid} deleted")

# ═══════════════════════════════════════════════════
# STOCK
# ═══════════════════════════════════════════════════
@app.get("/api/stock")
async def list_stock(q:Optional[str]=None, cat:Optional[str]=None,
                     conn=Depends(db), _=Depends(get_current_user)):
    where, params = [], []
    if q:
        params.append(f"%{q}%")
        where.append(f"(brand ILIKE ${len(params)} OR model ILIKE ${len(params)} OR spec ILIKE ${len(params)})")
    if cat: params.append(cat); where.append(f"category=${len(params)}")
    sql = "SELECT * FROM stock" + (" WHERE "+" AND ".join(where) if where else "") + " ORDER BY category,brand,model"
    rows = await conn.fetch(sql, *params)
    return [_out(r) for r in rows]

@app.post("/api/stock", status_code=201)
async def create_stock(s:StockCreate, conn=Depends(db), user=Depends(require_write)):
    row = await conn.fetchrow("""
        INSERT INTO stock (category,brand,model,spec,form_factor,interface,
            total_qty,avail_qty,alloc_qty,unit_cost,storage_loc,notes)
        VALUES ($1,$2,$3,$4,$5,$6,$7,$8,0,$9,$10,$11) RETURNING *""",
        s.cat,s.brand,s.model,s.spec,s.ff,s.ifc,s.total,s.avail,s.cost,s.loc,s.notes)
    item_name = f"{s.brand or ''} {s.model or ''}".strip() or s.cat
    await log_audit(conn, user, "STOCK_CREATED", "stock", str(row["id"]),
        f"Added stock: {item_name} | Cat: {s.cat} | Qty: {s.total} | Loc: {s.loc or '—'}")
    return _out(row)

@app.put("/api/stock/{sid}")
async def update_stock(sid:str, s:StockCreate, conn=Depends(db), user=Depends(require_write)):
    old = await conn.fetchrow("SELECT * FROM stock WHERE id=$1", sid)
    row = await conn.fetchrow("""
        UPDATE stock SET category=$1,brand=$2,model=$3,spec=$4,form_factor=$5,interface=$6,
            total_qty=$7,avail_qty=$8,unit_cost=$9,storage_loc=$10,notes=$11,updated_at=NOW()
        WHERE id=$12 RETURNING *""",
        s.cat,s.brand,s.model,s.spec,s.ff,s.ifc,s.total,s.avail,s.cost,s.loc,s.notes,sid)
    if not row: raise HTTPException(404,"Stock item not found")
    item_name = f"{s.brand or ''} {s.model or ''}".strip() or s.cat
    changes = []
    if old:
        if old["total_qty"] != s.total: changes.append(f"Qty: {old['total_qty']} → {s.total}")
        if old["avail_qty"] != s.avail: changes.append(f"Avail: {old['avail_qty']} → {s.avail}")
        if old["storage_loc"] != s.loc: changes.append(f"Loc: {old['storage_loc'] or '—'} → {s.loc or '—'}")
        if old["spec"] != s.spec: changes.append(f"Spec: {old['spec'] or '—'} → {s.spec or '—'}")
    detail = f"Updated stock: {item_name}"
    if changes: detail += " | " + " | ".join(changes)
    await log_audit(conn, user, "STOCK_UPDATED", "stock", sid, detail)
    return _out(row)

@app.delete("/api/stock/{sid}", status_code=204)
async def delete_stock(sid:str, conn=Depends(db), user=Depends(require_write)):
    old = await conn.fetchrow("SELECT category,brand,model FROM stock WHERE id=$1", sid)
    await conn.execute("DELETE FROM stock WHERE id=$1", sid)
    if old:
        item_name = f"{old['brand'] or ''} {old['model'] or ''}".strip() or old["category"]
        await log_audit(conn, user, "STOCK_DELETED", "stock", sid,
            f"Deleted stock item: {item_name} | Cat: {old['category']}")

@app.post("/api/stock/transaction")
async def stock_transaction(tx:StockTransaction, conn=Depends(db), user=Depends(require_write)):
    s = await conn.fetchrow("SELECT * FROM stock WHERE id=$1", tx.stock_id)
    if not s: raise HTTPException(404,"Stock item not found")
    avail,total,alloc = s["avail_qty"],s["total_qty"],s["alloc_qty"]
    if tx.tx_type=="IN":       total+=tx.qty; avail+=tx.qty
    elif tx.tx_type=="OUT":
        if avail<tx.qty: raise HTTPException(400,"Insufficient stock")
        avail-=tx.qty; total-=tx.qty
    elif tx.tx_type=="ALLOCATE":
        if avail<tx.qty: raise HTTPException(400,"Insufficient stock")
        avail-=tx.qty; alloc+=tx.qty
    elif tx.tx_type=="RETURN": avail+=tx.qty; alloc=max(0,alloc-tx.qty)
    elif tx.tx_type=="ADJUST": total=tx.qty; avail=tx.qty
    else: raise HTTPException(400,f"Unknown tx type: {tx.tx_type}")
    await conn.execute(
        "UPDATE stock SET total_qty=$1,avail_qty=$2,alloc_qty=$3,updated_at=NOW() WHERE id=$4",
        total,avail,alloc,tx.stock_id)
    await conn.execute(
        """INSERT INTO stock_transactions
           (stock_id,tx_type,qty,reference,notes,allocated_to,username)
           VALUES ($1,$2,$3,$4,$5,$6,$7)""",
        tx.stock_id, tx.tx_type, tx.qty, tx.ref, tx.notes,
        tx.allocated_to, user.get("username"))

    # ── Audit log ────────────────────────────────────────────────────
    item_name = f"{s['brand'] or ''} {s['model'] or ''}".strip() or s["category"]
    tx_labels = {"IN":"Received IN","OUT":"Issued OUT","ALLOCATE":"Allocated",
                 "RETURN":"Returned","ADJUST":"Stock Adjusted"}
    detail_parts = [
        f"{tx_labels.get(tx.tx_type,'Txn')} {tx.qty}x {item_name}",
        f"Category: {s['category']}",
        f"After: avail={avail} total={total} alloc={alloc}",
    ]
    if tx.allocated_to: detail_parts.append(f"Asset: {tx.allocated_to}")
    if tx.ref:          detail_parts.append(f"Ref: {tx.ref}")
    if tx.notes:        detail_parts.append(f"Notes: {tx.notes}")
    detail = " | ".join(detail_parts)

    # Primary audit: stock entity
    await log_audit(conn, user, f"STOCK_{tx.tx_type}", "stock", tx.stock_id,
        detail,
        related_entity="asset" if tx.allocated_to else None,
        related_entity_id=tx.allocated_to)

    # Secondary audit on the asset: appears in asset History tab
    if tx.allocated_to:
        asset_row = await conn.fetchrow(
            "SELECT id FROM assets WHERE hostname=$1", tx.allocated_to)
        if asset_row:
            asset_action = "ASSET_PART_ALLOCATED" if tx.tx_type in ("ALLOCATE","OUT") else "ASSET_PART_RETURNED"
            await log_audit(conn, user, asset_action, "asset", str(asset_row["id"]),
                detail,
                related_entity="stock",
                related_entity_id=tx.stock_id)

    return {"status":"ok","total":total,"avail":avail,"alloc":alloc}

@app.get("/api/stock/{sid}/transactions")
async def stock_txns(sid:str, conn=Depends(db), _=Depends(get_current_user)):
    rows = await conn.fetch(
        """SELECT * FROM stock_transactions
           WHERE stock_id=$1 ORDER BY created_at DESC LIMIT 200""", sid)
    return [_out(r) for r in rows]

# ═══════════════════════════════════════════════════
# CONNECTIVITY
# ═══════════════════════════════════════════════════
@app.get("/api/connectivity")
async def list_conns(q:Optional[str]=None, conn=Depends(db), _=Depends(get_current_user)):
    if q:
        rows = await conn.fetch("""SELECT * FROM connectivity WHERE
            src_hostname ILIKE $1 OR dst_hostname ILIKE $1 OR src_port_label ILIKE $1
            OR liu_a_hostname ILIKE $1 OR purpose ILIKE $1 ORDER BY src_hostname""", f"%{q}%")
    else:
        rows = await conn.fetch("SELECT * FROM connectivity ORDER BY src_hostname")
    return [_out(r) for r in rows]

@app.post("/api/connectivity", status_code=201)
async def create_conn(c:ConnCreate, conn=Depends(db), user=Depends(require_write)):
    row = await conn.fetchrow("""
        INSERT INTO connectivity (src_hostname,src_slot,src_port,src_port_label,
            liu_a_rack,liu_a_hostname,liu_a_port,liu_b_rack,liu_b_hostname,liu_b_port,
            dst_hostname,dst_port,cable_type,speed,vlan,purpose,notes)
        VALUES ($1,$2,$3,$4,$5,$6,$7,$8,$9,$10,$11,$12,$13,$14,$15,$16,$17) RETURNING *""",
        c.src_host,c.src_slot,c.src_port,c.src_label,c.liu_a_rack,c.liu_a_host,c.liu_a_port,
        c.liu_b_rack,c.liu_b_host,c.liu_b_port,c.dst_host,c.dst_port,
        c.cable,c.speed,c.vlan,c.purpose,c.notes)
    # Build readable path string
    path_parts = [p for p in [c.src_host, c.src_port or c.src_label,
                               c.liu_a_host, c.dst_host, c.dst_port] if p]
    detail = f"Connected: {' → '.join(path_parts)}"
    extras = [x for x in [c.cable, c.speed, f"VLAN {c.vlan}" if c.vlan else None, c.purpose] if x]
    if extras: detail += f" | {', '.join(extras)}"
    await log_audit(conn, user, "CONN_CREATED", "connectivity", str(row["id"]), detail,
        related_entity="asset" if c.src_host else None,
        related_entity_id=c.src_host)
    # Also log on the destination (switch) asset if it exists
    if c.dst_host:
        dst_asset = await conn.fetchrow("SELECT id FROM assets WHERE hostname=$1", c.dst_host)
        if dst_asset:
            await log_audit(conn, user, "CONN_CREATED", "connectivity", str(row["id"]), detail,
                related_entity="asset", related_entity_id=c.dst_host)
    return _out(row)

@app.put("/api/connectivity/{cid}")
async def update_conn(cid:str, c:ConnCreate, conn=Depends(db), user=Depends(require_write)):
    old = await conn.fetchrow("SELECT * FROM connectivity WHERE id=$1", cid)
    row = await conn.fetchrow("""
        UPDATE connectivity SET src_hostname=$1,src_slot=$2,src_port=$3,src_port_label=$4,
            liu_a_rack=$5,liu_a_hostname=$6,liu_a_port=$7,liu_b_rack=$8,liu_b_hostname=$9,
            liu_b_port=$10,dst_hostname=$11,dst_port=$12,cable_type=$13,speed=$14,
            vlan=$15,purpose=$16,notes=$17,updated_at=NOW()
        WHERE id=$18 RETURNING *""",
        c.src_host,c.src_slot,c.src_port,c.src_label,c.liu_a_rack,c.liu_a_host,c.liu_a_port,
        c.liu_b_rack,c.liu_b_host,c.liu_b_port,c.dst_host,c.dst_port,
        c.cable,c.speed,c.vlan,c.purpose,c.notes,cid)
    if not row: raise HTTPException(404,"Connection not found")
    # Detect meaningful changes
    changes = []
    if old:
        if old["dst_hostname"] != c.dst_host:
            changes.append(f"Switch: {old['dst_hostname'] or '—'} → {c.dst_host or '—'}")
        if old["dst_port"] != c.dst_port:
            changes.append(f"Port: {old['dst_port'] or '—'} → {c.dst_port or '—'}")
        if old["src_hostname"] != c.src_host:
            changes.append(f"Server: {old['src_hostname'] or '—'} → {c.src_host or '—'}")
        if old["cable_type"] != c.cable:
            changes.append(f"Cable: {old['cable_type'] or '—'} → {c.cable or '—'}")
        if old["speed"] != c.speed:
            changes.append(f"Speed: {old['speed'] or '—'} → {c.speed or '—'}")
        if old["vlan"] != c.vlan:
            changes.append(f"VLAN: {old['vlan'] or '—'} → {c.vlan or '—'}")
        if old["liu_a_hostname"] != c.liu_a_host:
            changes.append(f"LIU-A: {old['liu_a_hostname'] or '—'} → {c.liu_a_host or '—'}")
    if not changes: changes.append("Connectivity updated (no structural changes)")
    detail = f"Updated link {c.src_host or '?'} → {c.dst_host or '?'} | " + " | ".join(changes)
    await log_audit(conn, user, "CONN_UPDATED", "connectivity", cid, detail,
        related_entity="asset" if c.src_host else None,
        related_entity_id=c.src_host)
    return _out(row)

@app.delete("/api/connectivity/{cid}", status_code=204)
async def delete_conn(cid:str, conn=Depends(db), user=Depends(require_write)):
    old = await conn.fetchrow("SELECT * FROM connectivity WHERE id=$1", cid)
    await conn.execute("DELETE FROM connectivity WHERE id=$1", cid)
    if old:
        path_parts = [p for p in [old["src_hostname"], old["src_port"] or old["src_port_label"],
                                   old["liu_a_hostname"], old["dst_hostname"], old["dst_port"]] if p]
        detail = f"Removed: {' → '.join(path_parts)}"
        extras = [x for x in [old["cable_type"], old["speed"],
                               f"VLAN {old['vlan']}" if old["vlan"] else None] if x]
        if extras: detail += f" | {', '.join(extras)}"
        await log_audit(conn, user, "CONN_DELETED", "connectivity", cid, detail,
            related_entity="asset" if old["src_hostname"] else None,
            related_entity_id=old["src_hostname"])

# ═══════════════════════════════════════════════════
# USERS
# ═══════════════════════════════════════════════════
@app.get("/api/users")
async def list_users(conn=Depends(db), _=Depends(require_superuser)):
    rows = await conn.fetch(
        "SELECT id,name,username,role,email,dept,active,created_at,last_login FROM users ORDER BY name")
    return [_out(r) for r in rows]

@app.post("/api/users", status_code=201)
async def create_user(u:UserCreate, conn=Depends(db), _=Depends(require_superuser)):
    if not u.password: raise HTTPException(400,"Password required")
    ex = await conn.fetchval("SELECT id FROM users WHERE username=$1", u.username.lower())
    if ex: raise HTTPException(409,"Username already exists")
    row = await conn.fetchrow("""
        INSERT INTO users (name,username,pw_hash,role,email,dept)
        VALUES ($1,$2,$3,$4,$5,$6)
        RETURNING id,name,username,role,email,dept,active,created_at""",
        u.name,u.username.lower(),hash_pw(u.password),u.role,u.email,u.dept)
    return _out(row)

@app.put("/api/users/{uid}")
async def update_user(uid:str, u:UserCreate, conn=Depends(db), _=Depends(require_superuser)):
    ex = await conn.fetchval(
        "SELECT id FROM users WHERE username=$1 AND id!=$2", u.username.lower(), uid)
    if ex: raise HTTPException(409,"Username already exists")
    if u.password:
        row = await conn.fetchrow("""
            UPDATE users SET name=$1,username=$2,pw_hash=$3,role=$4,email=$5,dept=$6
            WHERE id=$7 RETURNING id,name,username,role,email,dept,active,created_at""",
            u.name,u.username.lower(),hash_pw(u.password),u.role,u.email,u.dept,uid)
    else:
        row = await conn.fetchrow("""
            UPDATE users SET name=$1,username=$2,role=$3,email=$4,dept=$5
            WHERE id=$6 RETURNING id,name,username,role,email,dept,active,created_at""",
            u.name,u.username.lower(),u.role,u.email,u.dept,uid)
    if not row: raise HTTPException(404,"User not found")
    return _out(row)

@app.delete("/api/users/{uid}", status_code=204)
async def delete_user(uid:str, conn=Depends(db), user=Depends(require_superuser)):
    if str(user["id"])==uid: raise HTTPException(400,"Cannot delete yourself")
    await conn.execute("DELETE FROM users WHERE id=$1", uid)

# ═══════════════════════════════════════════════════
# HW FIELDS
# ═══════════════════════════════════════════════════
@app.get("/api/hw-fields")
async def list_hw_fields(conn=Depends(db), _=Depends(get_current_user)):
    rows = await conn.fetch("SELECT * FROM hw_fields ORDER BY sort_order,created_at")
    return [_out(r) for r in rows]

@app.post("/api/hw-fields", status_code=201)
async def create_hw_field(f:HWField, conn=Depends(db), _=Depends(require_superuser)):
    ex = await conn.fetchval("SELECT id FROM hw_fields WHERE field_key=$1", f.key)
    if ex: raise HTTPException(409,"Field key already exists")
    row = await conn.fetchrow("""
        INSERT INTO hw_fields (field_key,label,field_type,placeholder,options,required,sort_order,is_system)
        VALUES ($1,$2,$3,$4,$5,$6,$7,false) RETURNING *""",
        f.key,f.label,f.field_type,f.placeholder,f.options,f.required,f.sort_order)
    return _out(row)


@app.put("/api/hw-fields/{fid}")
async def update_hw_field(fid:str, f:HWField, conn=Depends(db), _=Depends(require_superuser)):
    row = await conn.fetchrow("""
        UPDATE hw_fields SET label=$1,placeholder=$2,options=$3,required=$4
        WHERE id=$5 RETURNING *""",
        f.label, f.placeholder, f.options, f.required, fid)
    if not row: raise HTTPException(404,"Field not found")
    return _out(row)

@app.delete("/api/hw-fields/{fid}", status_code=204)
async def delete_hw_field(fid:str, conn=Depends(db), _=Depends(require_superuser)):
    sys = await conn.fetchval("SELECT is_system FROM hw_fields WHERE id=$1", fid)
    if sys: raise HTTPException(400,"Cannot delete system fields")
    await conn.execute("DELETE FROM hw_fields WHERE id=$1 AND is_system=false", fid)

# ═══════════════════════════════════════════════════
# EXPORT
# ═══════════════════════════════════════════════════
@app.get("/api/export/excel")
async def export_excel(conn=Depends(db), _=Depends(get_current_user)):
    wb  = openpyxl.Workbook()
    HDR = PatternFill("solid", fgColor="0D1424")
    HF  = Font(color="00C8FF", bold=True, name="Consolas", size=9)
    HA  = Alignment(horizontal="center", vertical="center")

    def style(ws, headers):
        ws.row_dimensions[1].height = 22
        for i,h in enumerate(headers,1):
            c = ws.cell(1,i,h)
            c.fill=HDR; c.font=HF; c.alignment=HA
            ws.column_dimensions[get_column_letter(i)].width = max(14,len(str(h))+4)

    def cv(v):
        """Sanitise a value so openpyxl can always write it safely.
        Handles: None, str, int, float, bool, datetime/date (tz-aware or naive),
        Decimal (PostgreSQL NUMERIC), asyncpg UUID/Range, and any other type."""
        if v is None:
            return ""
        if isinstance(v, bool):
            return str(v)
        if isinstance(v, int):
            return v
        if isinstance(v, float):
            return v
        if isinstance(v, str):
            return v
        # datetime / date — strip timezone info so openpyxl writes cleanly
        if hasattr(v, 'isoformat'):
            try:
                # Return naive datetime string openpyxl always accepts
                return v.strftime('%Y-%m-%d %H:%M:%S') if hasattr(v, 'hour') else v.strftime('%Y-%m-%d')
            except Exception:
                return str(v)
        # Decimal from PostgreSQL NUMERIC columns (e.g. unit_cost)
        try:
            import decimal
            if isinstance(v, decimal.Decimal):
                return float(v)
        except Exception:
            pass
        # asyncpg UUID, Record, Range, and any unknown type
        return str(v)

    try:
        hw_fields = await conn.fetch("SELECT field_key,label FROM hw_fields ORDER BY sort_order")

        # ── Assets sheet ──────────────────────────────────────────────
        ws_a = wb.active; ws_a.title="Assets"
        a_hdrs = ["hostname","asset_type","status",
                  "datacenter","rack_zone","rack_row","rack_id",
                  "u_start","u_height",
                  "prov_ip","bmc_ip","data_ip","bkup_ip","mac_addr","vlan",
                  "asset_tag","serial_number","po_number","eol_date","app_owner",
                  *[f["field_key"] for f in hw_fields],"notes"]
        style(ws_a, a_hdrs)

        rack_meta = {}
        for rk in await conn.fetch("SELECT rack_id,zone,row_label FROM racks"):
            rack_meta[str(rk["rack_id"])] = {
                "zone":      rk["zone"],
                "row_label": rk["row_label"],
            }

        for r in await conn.fetch("SELECT * FROM assets ORDER BY hostname"):
            hw = {}
            try:
                raw = r["hw_data"]
                if isinstance(raw, str):
                    hw = json.loads(raw)
                elif isinstance(raw, dict):
                    hw = raw
            except Exception:
                hw = {}
            rid      = r.get("rack_id") or ""
            rk       = rack_meta.get(str(rid), {})
            rack_zone = rk.get("zone") or ""
            rack_row  = rk.get("row_label") or ""
            row_vals = [
                cv(r.get("hostname")),   cv(r.get("asset_type")), cv(r.get("status")),
                cv(r.get("datacenter")), cv(rack_zone),           cv(rack_row),
                cv(rid),
                cv(r.get("u_start")),    cv(r.get("u_height")),
                cv(r.get("mgmt_ip")),    cv(r.get("oob_ip")),
                cv(hw.get("data_ip")),   cv(hw.get("bkup_ip")),
                cv(r.get("mac_addr")),   cv(r.get("vlan")),
                cv(r.get("asset_tag")),  cv(r.get("serial_number")),
                cv(hw.get("po_number")), cv(hw.get("eol_date")), cv(hw.get("app_owner")),
                *[cv(hw.get(f["field_key"])) for f in hw_fields],
                cv(r.get("notes"))
            ]
            ws_a.append(row_vals)

        # ── Stock sheet ───────────────────────────────────────────────
        ws_s = wb.create_sheet("Stock")
        style(ws_s,["category","brand","model","spec","form_factor","interface",
                    "total_qty","avail_qty","alloc_qty","storage_loc","unit_cost","notes"])
        for r in await conn.fetch("SELECT * FROM stock ORDER BY category,brand"):
            ws_s.append([cv(r.get("category")),cv(r.get("brand")),cv(r.get("model")),
                cv(r.get("spec")),cv(r.get("form_factor")),cv(r.get("interface")),
                cv(r.get("total_qty")),cv(r.get("avail_qty")),cv(r.get("alloc_qty")),
                cv(r.get("storage_loc")),cv(r.get("unit_cost")),cv(r.get("notes"))])

        # ── Racks sheet ───────────────────────────────────────────────
        ws_r = wb.create_sheet("Racks")
        style(ws_r,["rack_id","datacenter","zone","row_label","total_u","notes"])
        for r in await conn.fetch("SELECT * FROM racks ORDER BY datacenter,zone,row_label,rack_id"):
            ws_r.append([cv(r.get("rack_id")),cv(r.get("datacenter")),cv(r.get("zone")),
                         cv(r.get("row_label")),cv(r.get("total_u")),cv(r.get("notes"))])

        # ── Connectivity sheet ────────────────────────────────────────
        ws_c = wb.create_sheet("Connectivity")
        style(ws_c,["src_hostname","src_slot","src_port","src_port_label","liu_a_rack",
                    "liu_a_hostname","liu_a_port","liu_b_rack","liu_b_hostname","liu_b_port",
                    "dst_hostname","dst_port","cable_type","speed","vlan","purpose","notes"])
        for r in await conn.fetch("SELECT * FROM connectivity ORDER BY src_hostname"):
            ws_c.append([cv(r.get("src_hostname")),cv(r.get("src_slot")),cv(r.get("src_port")),
                cv(r.get("src_port_label")),cv(r.get("liu_a_rack")),cv(r.get("liu_a_hostname")),
                cv(r.get("liu_a_port")),cv(r.get("liu_b_rack")),cv(r.get("liu_b_hostname")),
                cv(r.get("liu_b_port")),cv(r.get("dst_hostname")),cv(r.get("dst_port")),
                cv(r.get("cable_type")),cv(r.get("speed")),cv(r.get("vlan")),
                cv(r.get("purpose")),cv(r.get("notes"))])

        # ── Audit Log sheet ───────────────────────────────────────────
        ws_al = wb.create_sheet("Audit Log")
        style(ws_al,["timestamp","user","action","entity","entity_id","detail"])
        for r in await conn.fetch(
            "SELECT * FROM audit_log ORDER BY created_at DESC LIMIT 5000"):
            ws_al.append([
                cv(r.get("created_at")), cv(r.get("username")),
                cv(r.get("action")),     cv(r.get("entity")),
                cv(r.get("entity_id")),  cv(r.get("detail"))
            ])

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        fname = f"dc-manager-{datetime.utcnow().strftime('%Y%m%d-%H%M%S')}.xlsx"
        return StreamingResponse(buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{fname}"'})

    except Exception as e:
        log.error(f"Export failed: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Export failed: {str(e)}")

# ═══════════════════════════════════════════════════
# IMPORT
# ═══════════════════════════════════════════════════
from fastapi import UploadFile, File

@app.post("/api/import/excel")
async def import_excel(sheet:str, file:UploadFile=File(...),
                       conn=Depends(db), user=Depends(require_write)):
    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
    ws_name = sheet if sheet in wb.sheetnames else \
              (sheet+"_Template" if sheet+"_Template" in wb.sheetnames else None)
    if not ws_name: raise HTTPException(400, f"Sheet '{sheet}' not found")
    rows = list(wb[ws_name].iter_rows(values_only=True))
    if len(rows)<2: return {"imported":0,"skipped":0,"errors":0}
    headers = [str(h).lower().strip() if h else "" for h in rows[0]]

    def col(row,key):
        try: i=headers.index(key); return row[i] if i<len(row) else None
        except: return None
    def s(v): return str(v).strip() if v is not None and str(v).strip() else None
    def n(v):
        try: return int(v) if v is not None else None
        except: return None

    hw_fields = await conn.fetch("SELECT field_key FROM hw_fields")
    hw_keys   = [f["field_key"] for f in hw_fields]
    imported = errors = skipped = 0

    for row in rows[1:]:
        if not any(row): continue
        try:

            # ── ASSETS ──────────────────────────────────────────────
            if sheet=="Assets":
                host = s(col(row,"hostname"))
                if not host: continue

                # Build hw_data from custom fields + named extra fields
                hw = {k:s(col(row,k)) for k in hw_keys if s(col(row,k))}
                for fk in ["data_ip","bkup_ip","po_number","eol_date","app_owner"]:
                    v = s(col(row,fk))
                    if v: hw[fk] = v

                prov_ip = s(col(row,"prov_ip")) or s(col(row,"mgmt_ip"))
                bmc_ip  = s(col(row,"bmc_ip"))  or s(col(row,"oob_ip"))
                rack_id = s(col(row,"rack_id"))
                dc      = s(col(row,"datacenter"))
                r_zone  = s(col(row,"rack_zone")) or s(col(row,"zone"))
                r_row   = s(col(row,"rack_row"))  or s(col(row,"row_label"))
                u_st    = n(col(row,"u_start"))
                u_ht    = n(col(row,"u_height")) or 1
                a_type  = s(col(row,"asset_type")) or "Server"
                a_stat  = s(col(row,"status"))     or "Online"

                # Auto-upsert rack record if zone/row/dc provided
                if rack_id and (r_zone or r_row or dc):
                    await conn.execute("""
                        INSERT INTO racks (rack_id,datacenter,zone,row_label,total_u)
                        VALUES ($1,$2,$3,$4,42)
                        ON CONFLICT (rack_id) DO UPDATE
                        SET datacenter=COALESCE($2, racks.datacenter),
                            zone=COALESCE($3, racks.zone),
                            row_label=COALESCE($4, racks.row_label)""",
                        rack_id, dc, r_zone, r_row)

                ex = await conn.fetchval("SELECT id FROM assets WHERE hostname=$1", host)
                if ex:
                    # Hostname already exists — skip (import = new records only)
                    skipped += 1
                    log.info(f"Import skip asset '{host}': already exists (id={ex})")
                    continue

                # Slot conflict check
                if rack_id and u_st:
                    u_end = u_st + u_ht - 1
                    conflict = await conn.fetchrow("""
                        SELECT hostname,u_start,u_height FROM assets
                        WHERE rack_id=$1
                          AND u_start IS NOT NULL
                          AND (u_start <= $3 AND u_start + u_height - 1 >= $2)""",
                        rack_id, u_st, u_end)
                    if conflict:
                        log.warning(f"Import skip '{host}': slot conflict in {rack_id} "
                                    f"U{u_st}-{u_end} with {conflict['hostname']}")
                        errors += 1
                        continue

                new_row = await conn.fetchrow("""INSERT INTO assets
                    (hostname,asset_type,status,server_type,datacenter,rack_id,
                     u_start,u_height,mgmt_ip,oob_ip,mac_addr,vlan,
                     asset_tag,serial_number,notes,hw_data)
                    VALUES ($1,$2,$3,$4,$5,$6,$7,$8,$9,$10,$11,$12,$13,$14,$15,$16)
                    RETURNING id""",
                    host, a_type, a_stat, s(col(row,"server_type")),
                    dc, rack_id, u_st, u_ht,
                    prov_ip, bmc_ip, s(col(row,"mac_addr")), s(col(row,"vlan")),
                    s(col(row,"asset_tag")), s(col(row,"serial_number")),
                    s(col(row,"notes")), json.dumps(hw))

                await log_audit(conn, user, "ASSET_CREATED", "asset", str(new_row["id"]),
                    f"Imported: {host} | Type: {a_type} | Status: {a_stat} "
                    f"| DC: {dc or '—'} | Rack: {rack_id or '—'} | U: {u_st or '—'}")
                imported += 1

            # ── STOCK ────────────────────────────────────────────────
            elif sheet=="Stock":
                cat   = s(col(row,"category"))
                brand = s(col(row,"brand"))
                model = s(col(row,"model"))
                if not cat: continue

                # Skip if exact same brand+model already exists
                if brand and model:
                    ex = await conn.fetchval(
                        "SELECT id FROM stock WHERE brand=$1 AND model=$2", brand, model)
                    if ex:
                        skipped += 1
                        log.info(f"Import skip stock '{brand} {model}': already exists")
                        continue

                total = n(col(row,"total_qty")) or 0
                avail = n(col(row,"avail_qty")) or 0
                alloc = n(col(row,"alloc_qty")) or 0
                # avail cannot exceed total
                avail = min(avail, total)

                new_row = await conn.fetchrow("""INSERT INTO stock
                    (category,brand,model,spec,form_factor,interface,
                     total_qty,avail_qty,alloc_qty,unit_cost,storage_loc,notes)
                    VALUES ($1,$2,$3,$4,$5,$6,$7,$8,$9,$10,$11,$12)
                    RETURNING id""",
                    cat, brand, model, s(col(row,"spec")),
                    s(col(row,"form_factor")), s(col(row,"interface")),
                    total, avail, alloc,
                    float(col(row,"unit_cost")) if col(row,"unit_cost") else None,
                    s(col(row,"storage_loc")), s(col(row,"notes")))

                item_name = f"{brand or ''} {model or ''}".strip() or cat
                await log_audit(conn, user, "STOCK_CREATED", "stock", str(new_row["id"]),
                    f"Imported: {item_name} | Cat: {cat} | Total: {total} | Avail: {avail}")
                imported += 1

            # ── RACKS ────────────────────────────────────────────────
            elif sheet=="Racks":
                rid = s(col(row,"rack_id"))
                if not rid: continue
                await conn.execute("""
                    INSERT INTO racks (rack_id,datacenter,zone,row_label,total_u,notes)
                    VALUES ($1,$2,$3,$4,$5,$6)
                    ON CONFLICT (rack_id) DO UPDATE
                    SET datacenter=$2, zone=$3, row_label=$4, total_u=$5, notes=$6""",
                    rid, s(col(row,"datacenter")), s(col(row,"zone")),
                    s(col(row,"row_label")), n(col(row,"total_u")) or 42,
                    s(col(row,"notes")))
                await log_audit(conn, user, "RACK_CREATED", "rack", rid,
                    f"Imported rack: {rid} | DC: {s(col(row,'datacenter')) or '—'}")
                imported += 1

            # ── CONNECTIVITY ─────────────────────────────────────────
            elif sheet=="Connectivity":
                src      = s(col(row,"src_hostname"))
                src_port = s(col(row,"src_port")) or s(col(row,"src_port_label"))
                dst      = s(col(row,"dst_hostname"))
                if not src: continue

                # Duplicate check — a server can have many connections (one per port/label)
                # Use src_port_label as the unique key per connection since it encodes
                # slot+port together (e.g. "slot1port1", "slot2port1" are different links)
                src_port_label = s(col(row,"src_port_label"))
                if src_port_label:
                    ex = await conn.fetchval(
                        "SELECT id FROM connectivity WHERE src_hostname=$1 AND src_port_label=$2",
                        src, src_port_label)
                    if ex:
                        skipped += 1
                        log.info(f"Import skip connectivity {src}/{src_port_label}: already exists")
                        continue
                elif src_port and dst:
                    # No label — fall back to src+port+dst as composite key
                    ex = await conn.fetchval(
                        """SELECT id FROM connectivity
                           WHERE src_hostname=$1 AND src_port=$2 AND dst_hostname=$3""",
                        src, src_port, dst)
                    if ex:
                        skipped += 1
                        log.info(f"Import skip connectivity {src}/{src_port}->{dst}: already exists")
                        continue

                new_row = await conn.fetchrow("""INSERT INTO connectivity
                    (src_hostname,src_slot,src_port,src_port_label,
                     liu_a_rack,liu_a_hostname,liu_a_port,
                     liu_b_rack,liu_b_hostname,liu_b_port,
                     dst_hostname,dst_port,cable_type,speed,vlan,purpose,notes)
                    VALUES ($1,$2,$3,$4,$5,$6,$7,$8,$9,$10,$11,$12,$13,$14,$15,$16,$17)
                    RETURNING id""",
                    src, s(col(row,"src_slot")), s(col(row,"src_port")),
                    s(col(row,"src_port_label")),
                    s(col(row,"liu_a_rack")), s(col(row,"liu_a_hostname")),
                    s(col(row,"liu_a_port")),
                    s(col(row,"liu_b_rack")), s(col(row,"liu_b_hostname")),
                    s(col(row,"liu_b_port")),
                    dst, s(col(row,"dst_port")), s(col(row,"cable_type")),
                    s(col(row,"speed")), s(col(row,"vlan")),
                    s(col(row,"purpose")), s(col(row,"notes")))

                path_parts = [p for p in [src, src_port, s(col(row,"liu_a_hostname")), dst,
                                          s(col(row,"dst_port"))] if p]
                detail = f"Imported: {' → '.join(path_parts)}"
                extras = [x for x in [s(col(row,"cable_type")), s(col(row,"speed")),
                          f"VLAN {s(col(row,'vlan'))}" if s(col(row,"vlan")) else None] if x]
                if extras: detail += f" | {', '.join(extras)}"
                await log_audit(conn, user, "CONN_CREATED", "connectivity",
                    str(new_row["id"]), detail,
                    related_entity="asset" if src else None,
                    related_entity_id=src)
                imported += 1

        except Exception as e:
            log.warning(f"Import row error ({sheet}): {e}")
            errors += 1

    log.info(f"Import {sheet}: imported={imported} skipped={skipped} errors={errors}")
    return {"imported": imported, "skipped": skipped, "errors": errors}
